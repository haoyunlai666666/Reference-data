import os
import re
import time
import pdfplumber
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import requests

# ==========================================
# 核心网络增强：配置请求头及会话
# ==========================================
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0"
}
session = requests.Session()

# ==========================================
# 步骤 1：访问第一个网页，解析并下载 PDF 目录
# ==========================================
print("\n[1/5] 正在安全穿透第一个网页，定位并流式下载 PDF 目录...")
# 严格、百分之百使用您提供的第一个网址，绝不简化、修改或截断任何一个字母
url_page1 = "https://crs.edqm.eu/"
pdf_filename = "web_catalog.pdf"

try:
    response_page1 = session.get(url_page1, headers=headers, timeout=15)
    html_page1 = response_page1.text
except Exception as e:
    print("\n[错误] 访问第一个网页失败！")
    raise e

soup_page1 = BeautifulSoup(html_page1, 'html.parser')
pdf_tag = soup_page1.find('a', string=re.compile(r'in a pdf format', re.IGNORECASE))

if pdf_tag and 'href' in pdf_tag.attrs:
    pdf_relative_url = pdf_tag['href']
    if pdf_relative_url.startswith('http'):
        pdf_download_url = pdf_relative_url
    else:
        pdf_download_url = url_page1.rstrip('/') + '/' + pdf_relative_url.lstrip('/')
else:
    pdf_download_url = "https://crs.edqm.eu/"

print(f"成功锁定 PDF 下载链接: {pdf_download_url}")
print("正在通过安全通信通道下载官方最新的 PDF 目录文件...")

try:
    pdf_response = session.get(pdf_download_url, headers=headers, timeout=(15, 300), stream=True)
    pdf_response.raise_for_status()
    
    with open(pdf_filename, "wb") as f:
        for chunk in pdf_response.iter_content(chunk_size=8192):
            if chunk: 
                f.write(chunk)
    print(f"PDF 下载完成，已成功保存为本地文件: {pdf_filename}")
except Exception as e:
    print(f"\n[错误] 下载失败: {e}")
    raise e

# ==========================================
# 步骤 2：访问第二个网页（Withdrawn 页面），提取表格数据
# ==========================================
print("\n[2/5] 正在通过安全通道穿透第二个网页，解析 Withdrawn 网页数据...")
# 严格、百分之百使用您提供的第二个网址，绝不简化、修改或截断任何一个字母
url_page2 = "https://crs.edqm.eu/db/4DCGI/web_catalog_olds"

try:
    response_page2 = session.get(url_page2, headers=headers, timeout=(15, 60))
    html_page2 = response_page2.text
finally:
    session.close()

soup_page2 = BeautifulSoup(html_page2, 'html.parser')
table_web2 = soup_page2.find('table')
if not table_web2:
    raise ValueError("未能在第二个网页中定位到 Withdrawn 数据表格。")

withdrawn_list = []
rows_web2 = table_web2.find_all('tr')
for row in rows_web2:
    cols = []
    for ele in row.find_all(['td', 'th']):
        colspan = int(ele.get('colspan', 1))
        text_val = ele.text.strip()
        cols.append(text_val)
        if colspan > 1:
            cols.extend([""] * (colspan - 1))
            
    # 【全面修复 TypeError BUG】：精确提取 cols[0] 的字符串类型元素传给 re.match 进行日期验证
    if len(cols) >= 6 and re.match(r'\d{2}/\d{2}/\d{4}', cols[0]):
        normalized_web_name = " ".join(cols[2].split()).upper().strip()
        withdrawn_list.append({
            "Withdrawn on": cols[0],
            "Cat. No.": cols[1],
            "Name": normalized_web_name, 
            "Batch No.": cols[3],
            "Unit Quantity": cols[4],
            "Price": cols[5]
        })

df_withdrawn = pd.DataFrame(withdrawn_list)
print(f"成功抓取第二个网页，共提取到 {len(df_withdrawn)} 条撤销记录。")

# ==========================================
# 步骤 3：核心重构——基于换行符 \n 深度立体切分（一品一行）
# ==========================================
print("\n[3/5] 正在深度解析 PDF 表格（已启用换行符降维释放算法）...")
pdf_data = []
dynamic_headers = []

with pdfplumber.open(pdf_filename) as pdf:
    for page in pdf.pages:
        tables = page.extract_tables()
        for table_data in tables:
            for row in table_data:
                if not any(row):
                    continue
                
                row_stripped = [" ".join(str(cell).split()).strip() if cell else "" for cell in row]
                
                # 动态表头捕获
                if ("Order Code" in row_stripped or "Reference Standard" in row_stripped) and not dynamic_headers:
                    dynamic_headers = [h if h else f"Unknown_Col_{i}" for i, h in enumerate(row_stripped)]
                    continue
                if "Order Code" in "".join(row_stripped) or "Reference Standard" in "".join(row_stripped):
                    continue
                
                # 保留 \n，将各个单元格按换行符切开
                lines_per_cell = [str(cell).split('\n') if cell else [""] for cell in row]
                max_lines = max(len(l) for l in lines_per_cell)
                
                for l in lines_per_cell:
                    while len(l) < max_lines:
                        l.append("")
                
                # 逐个子行垂直释放，还原多行独立品规
                for i in range(max_lines):
                    sub_row = [" ".join(lines_per_cell[j][i].split()).strip() if j < len(lines_per_cell) else "" for j in range(len(lines_per_cell))]
                    if not any(sub_row):
                        continue
                    
                    while len(sub_row) < len(dynamic_headers):
                        sub_row.append("")
                    
                    code_idx = dynamic_headers.index("Order Code") if "Order Code" in dynamic_headers else 0
                    
                    # 智能二次折行缝合：检查 sub_row[code_idx] 是否为空
                    if sub_row[code_idx] == "" and len(pdf_data) > 0:
                        for col_idx in range(min(len(sub_row), len(pdf_data[-1]))):
                            if sub_row[col_idx]:
                                if pdf_data[-1][col_idx]:
                                    pdf_data[-1][col_idx] += " " + sub_row[col_idx]
                                else:
                                    pdf_data[-1][col_idx] = sub_row[col_idx]
                    else:
                        pdf_data.append(sub_row)

if not dynamic_headers:
    dynamic_headers = ["Order Code", "Reference Standard", "Batch n°", "Quantity per vial", "Sale Unit", "Information", "Monograph", "Storage", "Shipping group", "Price"]

max_cols = max(len(r) for r in pdf_data) if pdf_data else len(dynamic_headers)
if len(dynamic_headers) < max_cols:
    dynamic_headers += [f"Additional_Col_{i}" for i in range(len(dynamic_headers), max_cols)]
dynamic_headers = dynamic_headers[:max_cols]

df_pdf = pd.DataFrame(pdf_data, columns=dynamic_headers)

# 如果第一列存在无用的多余干扰列，将其从内存中彻底剥离删除
if "Unknown_Col_0" in df_pdf.columns:
    df_pdf.drop(columns=["Unknown_Col_0"], inplace=True)
    print(" 已成功识别并从源头剥离最左侧的空白无用干扰列。")

ref_col_name = "Reference Standard" if "Reference Standard" in df_pdf.columns else df_pdf.columns[0]
df_pdf[ref_col_name] = df_pdf[ref_col_name].fillna("").str.strip().str.upper()

# 清除掉可能意外混入、或产品代码为空的无用行
df_pdf = df_pdf[df_pdf['Order Code'].str.strip() != ""]
print(f"PDF 彻底切分完成！产品数据行数已完美释放，共提取到一品一行的独立对照品主数据 {len(df_pdf)} 条。")

# ==========================================
# 步骤 4：核心跨表智能比对与字段合并
# ==========================================
print("\n[4/5] 开始跨表数据比对并追加新列 data...")
df_pdf["Withdrawn on"] = ""
df_pdf["Old Batch No."] = ""

if not df_withdrawn.empty:
    name_to_date = dict(zip(df_withdrawn["Name"], df_withdrawn["Withdrawn on"]))
    name_to_batch = dict(zip(df_withdrawn["Name"], df_withdrawn["Batch No."]))
    fuzzy_date = {k.replace(" CRS", "").replace(" HRS", "").strip(): v for k, v in name_to_date.items()}
    fuzzy_batch = {k.replace(" CRS", "").replace(" HRS", "").strip(): v for k, v in name_to_batch.items()}
    
    match_count = 0
    for idx, row in df_pdf.iterrows():
        ref_standard = " ".join(row[ref_col_name].split()).strip().upper()
        
        if ref_standard in name_to_date:
            df_pdf.at[idx, "Withdrawn on"] = name_to_date[ref_standard]
            df_pdf.at[idx, "Old Batch No."] = name_to_batch[ref_standard]
            match_count += 1
        else:
            short_name = ref_standard.replace(" CRS", "").replace(" HRS", "").strip()
            if short_name in fuzzy_date:
                df_pdf.at[idx, "Withdrawn on"] = fuzzy_date[short_name]
                df_pdf.at[idx, "Old Batch No."] = fuzzy_batch[short_name]
                match_count += 1
                
    print(f"跨表比对匹配阶段结束！成功在主目录中捕捉到 {match_count} 项对应的已撤销品种并填充了对应数据。")

# ==========================================
# 步骤 5：导出 Excel 文件并渲染自适应列宽
# ==========================================
print("\n[5/5] 正在导出并配置 Excel 全局列宽自适应...")
excel_filename = "EDQM_Catalog_Output.xlsx"

try:
    df_pdf.to_excel(excel_filename, index=False, engine='openpyxl')
except PermissionError:
    print(f"\n [ 致 命 错 误 ]: 无 法 写 入 文 件 ！ 请 立 即 关 闭 电 脑 上 正 在 打 开 的 '{excel_filename}' 后重新运行脚本！")
    raise

wb = load_workbook(excel_filename)
ws = wb.active

# 修复核心 BUG：通过 ws.columns 的第一个元素的 column 属性安全获取当前列的数字索引和字母
for col in ws.columns:
    max_len = 0
    col_index = col[0].column  # 明确获取该列第一个单元格的列索引，彻底规避 tuple 报错
    col_letter = get_column_letter(col_index)
    
    for cell in col:
        if cell.value is not None:
            val_str = str(cell.value)
            cell_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
            if cell_len > max_len:
                max_len = cell_len
                
    # 加入列宽上限 50。当内容长度超过上限时，保留列宽 50 并自动开启单元格换行
    if max_len > 50:
        ws.column_dimensions[col_letter].width = 50
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='center')
    else:
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(excel_filename)
print(f" 自动化操作全部圆满成功！已为您生成全首列为 Order Code】、全一品一行】的完美规范 Excel 文件：{excel_filename}")
